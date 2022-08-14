#https://figshare.com/articles/dataset/Virtual_E_Dataset/5513407
#Han Seung Seog (whria78@gmail.com)
#https://modelderm.com

def main():
    import os
    import torch
    from torchvision import datasets, transforms

    import numpy as np
    import argparse
    import time
    import timm

    #parse arguments
    parser = argparse.ArgumentParser(description='An example of CNN training and deployment; Han Seung Seog')
    parser.add_argument('--model', type=str, default='mobilenet', help='mobilenet / efficientnet / vgg (mobilenet by default)')
    parser.add_argument('--resolution', type=int, default=224, help='image resolution (224 by default)')

    parser.add_argument('--epoch', type=int, default=6, help='number of epochs to train (6 by default)')
    parser.add_argument('--batch', type=int, default=32, help='batch size (32 by default)')

    parser.add_argument('--lr', type=float, default=0.01, help='learning rate (0.005 by default)')
    parser.add_argument('--gamma', type=float, default=0.1, help='gamma (0.005 by default)')
    parser.add_argument('--step', type=int, default=3, help='step (learning rate = learning rate * gamma; every of 3 epochs by default)')

    parser.add_argument('--train', type=str, default='dataset/train', help='training image folder (/dataset/train by default)')
    parser.add_argument('--val', type=str, default='dataset/val', help='validation image folder (/dataset/val by default)')
    parser.add_argument('--test', type=str, default='dataset/test', help='test image folder (/dataset/test by default)')

    args = parser.parse_args()

    #define transformations
    dataset_means=(0.485, 0.456, 0.406) #precomputed channel means of ImageNet(train) for normalization
    dataset_stds=(0.229, 0.224, 0.225) #precomputed standard deviations

    transformations = {
        'train': transforms.Compose([
            transforms.Resize((args.resolution,args.resolution)),
            transforms.RandomHorizontalFlip(),
            transforms.ToTensor(),
            transforms.Normalize(dataset_means, dataset_stds)
        ]),
        'val': transforms.Compose([
            transforms.Resize((args.resolution,args.resolution)),
            transforms.ToTensor(),
            transforms.Normalize(dataset_means, dataset_stds)
        ])
    }

    #load datasets and define loaders

    imagedir_train=os.path.join(os.getcwd(),args.train)
    imagedir_val=os.path.join(os.getcwd(),args.val)
    imagedir_test=os.path.join(os.getcwd(),args.test)

    def print_dataset_info(dataset_):
        _, list_dx_count_list = torch.unique(torch.tensor(dataset_.targets), return_counts=True)
        list_dx_count_list=list_dx_count_list.to("cpu").numpy().tolist()
        for class_no,class_name in enumerate(dataset_.class_to_idx):
            print(class_name,list_dx_count_list[class_no], 'images')

    train_dataset=datasets.ImageFolder(imagedir_train,transformations['train'])
    print("Train : ",imagedir_train)
    print_dataset_info(train_dataset)
    val_dataset=datasets.ImageFolder(imagedir_val,transformations['val'])
    print("Validation : ",imagedir_val)
    print_dataset_info(val_dataset)
    test_dataset=datasets.ImageFolder(imagedir_test,transformations['val'])
    print("Test : ",imagedir_test)
    print_dataset_info(test_dataset)


    train_loader = torch.utils.data.DataLoader(train_dataset, batch_size=args.batch, shuffle=True, pin_memory=True, drop_last=True,num_workers=1)
    val_loader = torch.utils.data.DataLoader(val_dataset, batch_size=args.batch, shuffle=False, pin_memory=True, drop_last=True,num_workers=1)
    test_loader = torch.utils.data.DataLoader(test_dataset, batch_size=args.batch, shuffle=False, pin_memory=True, drop_last=True,num_workers=1)

    if train_dataset.class_to_idx != val_dataset.class_to_idx:
        print("ERROR - TRAIN-VAL : ",list(set(train_dataset.classes)-set(val_dataset.classes)))
        print("ERROR - VAL-TRAIN : ",list(set(val_dataset.classes)-set(train_dataset.classes)))
        sys.exit(1)
    len_classes=len(train_dataset.class_to_idx)

    #save class info for DEMO
    save_path=os.path.join(os.getcwd(),'demo','dxinfo.js')
    try:os.makedirs(os.path.dirname(save_path))
    except:pass
    
    save_string="list_dx=["
    for classname in train_dataset.classes:
        save_string+='"%s",' % (classname)
    save_string=save_string[:-1]
    save_string+="];"

    f=open(save_path,'w')
    f.write(save_string)
    f.close()
    print("List of disease classes is saved to : ", save_path)


    #choose model to train
    if args.model == 'mobilenet':
        model=timm.create_model('mobilenetv2_100',num_classes=len_classes,pretrained=True)
    elif args.model == 'efficientnet':
        model=timm.create_model('efficientnet_lite0',num_classes=len_classes,pretrained=True)
    elif args.model == 'vgg':
        model=timm.create_model('vgg19_bn',num_classes=len_classes,pretrained=True)
    else:
        print("ERR - No dataset")

    #number of iterations
    n_epochs = args.epoch
    #optimizer and scheduler
    optimizer = torch.optim.SGD(model.parameters(), lr=args.lr, momentum=0.9)
    scheduler = torch.optim.lr_scheduler.MultiStepLR(optimizer, milestones=[args.step, args.step*2, args.step*3], gamma=args.gamma)
    #loss function (standard cross-entropy taking logits as inputs)
    loss_fn = torch.nn.CrossEntropyLoss()
    #train on GPU if CUDA is available, else on CPU
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    model = model.to(device)


    #print training information
    print("")
    if torch.cuda.is_available():
        hardware = "GPU " + str(device) 
    else:
        hardware = "CPU (CUDA was not found)" 
    print("Training information:")
    print("hardware:", hardware)
    print("total number of epochs:", n_epochs)
    print("mini batch size:", args.batch)
    print("")

    #training loop
    lowest_val_loss = np.inf #used for saving the best model with lowest validation loss
    for epoch in range(1, n_epochs+1):
        #train model
        start_time = time.time()
        train_losses = []
        model.train()
        for i, (imgs, labels) in enumerate(train_loader):
            batch_size = imgs.shape[0]
            print("train mini batch " + str(i+1) + "/" + str(len(train_loader)) + " - %d training images processed" % (i*batch_size), end="\r", flush=True) 
            imgs, labels = imgs.to(device), labels.to(device)
            outputs = model(imgs)
            loss = loss_fn(outputs, labels)
            optimizer.zero_grad()
            loss.backward()
            optimizer.step()
            train_losses.append(loss.item())

        scheduler.step()
        
        print("                                                                         ", end="\r", flush=True) #delete output from train counter to not interfere with validation counter (probably can be done better)

        #validate model
        with torch.no_grad():
            model.eval()
            correct_labels = 0
            all_labels = 0
            val_losses = []
            for i, (imgs, labels) in enumerate(val_loader):
                print("valid batch " + str(i+1) + "/" + str(len(val_loader)), end="\r", flush=True) 
                imgs, labels = imgs.to(device), labels.to(device)
                batch_size = imgs.shape[0]
                outputs = model(imgs)
                loss = loss_fn(outputs, labels)
                val_losses.append(loss.item())
                _, preds = torch.max(outputs, dim=1) #predictions
                matched = preds == labels #comparison with ground truth
                
                correct_labels += float(torch.sum(matched)) 
                all_labels += float(batch_size) 

            val_accuracy = correct_labels / all_labels #compute top-1 accuracy on validation data 
        
        train_loss = np.mean(train_losses)
        val_loss = np.mean(val_losses)

        #save best model so far
        save_path=os.path.join(os.getcwd(),'trained_model',args.model + '.pth')
        try:os.makedirs(os.path.dirname(save_path))
        except:pass
        torch.save(model.state_dict(), save_path)
        torch_save_path=save_path
        
        end_time = time.time()
        
        #print iteration results
        print("Epoch: %d/%d, lr: %f, train_loss: %f, val_loss: %f, val_acc: %f, time(sec): %f" % (epoch, n_epochs, optimizer.param_groups[0]['lr'], train_loss, val_loss, val_accuracy, end_time - start_time))

    print("Save Pytorch model to : ",torch_save_path)
    print("Finish Training")

    #load the last model
    #save_path=os.path.join(os.getcwd(),'trained_model',args.model + '.pth')
    #model.load_state_dict(torch.load(save_path))


    #TEST
    start_time = time.time()

    all_label_list=[] #for auroc
    all_output_list=[] #for auroc

    def softmax(x):
        """Compute softmax values for each sets of scores in x."""
        e_x = np.exp(x - np.max(x))
        return e_x / e_x.sum(axis=0) # only difference   

    with torch.no_grad():
        model.eval()
        correct_labels = 0
        all_labels = 0
        val_losses = []
        for i, (imgs, labels) in enumerate(test_loader):
            print("test batch " + str(i+1) + "/" + str(len(test_loader)), end="\r", flush=True) 
            imgs, labels = imgs.to(device), labels.to(device)
            batch_size = imgs.shape[0]
            outputs = model(imgs)
            loss = loss_fn(outputs, labels)
            val_losses.append(loss.item())
            _, preds = torch.max(outputs, dim=1) #predictions
            matched = preds == labels #comparison with ground truth
            correct_labels += float(torch.sum(matched))
            all_labels += float(batch_size)
            
            all_label_list+=labels.to("cpu").numpy().tolist()
            all_output_list+=outputs.to("cpu").numpy().tolist()

        test_accuracy = correct_labels / all_labels #compute top-1 accuracy on validation data 
        end_time = time.time()
        print("TEST Accuracy: %f, time(sec): %f" % (test_accuracy, end_time - start_time))

    #Calculate AUC save xls for R statistics
    from sklearn.metrics import roc_auc_score
    from openpyxl import Workbook
    
    for class_no,class_name in enumerate(train_dataset.class_to_idx):
        print(class_name)
        y_real=[]
        y_pred=[]
        for no_,all_output_list_ in enumerate(all_output_list): 
            all_output_list_softmax_=softmax(all_output_list_)
            if  all_label_list[no_]==class_no:
                y_real+=[1]
            else:
                y_real+=[0]
            y_pred+=[all_output_list_softmax_[class_no]]
  
        score = roc_auc_score(np.array(y_real), np.array(y_pred))
        print(f"ROC AUC: {score:.4f}")

        wb = Workbook()    
        sheet1 = wb.active
        sheet1.title = class_name
        sheet1.cell(row=1, column=1).value = "dx"
        sheet1.cell(row=1, column=2).value = "pred"
        for no_,all_output_list_ in enumerate(all_output_list): 
            all_output_list_softmax_=softmax(all_output_list_)
            sheet1.cell(row=(no_+2), column=1).value = (all_label_list[no_]==class_no)
            sheet1.cell(row=(no_+2), column=2).value = all_output_list_softmax_[class_no]

        save_path=os.path.join(os.getcwd(),'stat','roc_%s.xlsx' % (class_name))
        try:os.makedirs(os.path.dirname(save_path))
        except:pass
        wb.save(filename=save_path)
        excel_path=save_path

        f=open(os.path.join(os.getcwd(),'template','r.template'),'r')
        template_list=f.readlines()
        f.close()
        template_=''
        for t_ in template_list:
            template_+=t_+'\n'
        
        template_=template_.replace('[XLS_PATH]',excel_path).replace('[TEST_NO_IMAGES]','%d' % (len(test_dataset))).replace('\\','/')
        
        save_path=os.path.join(os.getcwd(),'stat','R_%s.txt' % (class_name))
        f=open(save_path,'w')
        f.write(template_)
        f.close()
        

    model.eval()
    save_path=os.path.join(os.getcwd(),'demo','model.onnx')
    try:os.makedirs(os.path.dirname(save_path))
    except:pass

    torch.onnx.export(model, torch.autograd.Variable(torch.randn(1, 3, args.resolution, args.resolution)), save_path, verbose=False,opset_version=10)
    print("Export the ONNX model to : ",save_path)
    
        
if __name__ == '__main__':
    main()
